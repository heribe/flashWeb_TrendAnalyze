# -*- coding: utf-8 -*-
import numpy as np
import matplotlib
# matplotlib.use('TkAgg') #这行是对mac上matplotlib无法使用的解决方式
# matplotlib.use('Agg') # 这行是解决linux上画图问题
import matplotlib.pyplot as plt
import logging
# import mysql.connector
from openpyxl import load_workbook
import time

## debug，info，warning，error
logging.basicConfig(level=logging.ERROR)


## 画图
def graphIt(data, w, s):
    x = data[:, 0]
    y = data[:, 1]
    ## 折线图
    plt.plot(x, y, marker='o', label="data")

    plt.plot([1, len(data)], [w[0] + w[1], w[0] + w[1] * len(data)], label="trend")
    plt.title(s)
    plt.xlabel("month")
    plt.ylabel("cases")
    plt.legend()
    plt.show()


def caseTrend(data, s):
    x = data[:, 0]
    y = data[:, 1]

    x_ = np.zeros((len(data), 2))
    x_[:, 1] = data[:, 0]
    x_[:, 0] = np.ones(len(x))
    w = np.linalg.inv(x_.T.dot(x_)).dot(x_.T).dot(y)

    return w


## 计算标准差
def calE(x, y, w):
    Ein = np.array([])
    for i in range(len(x)):
        Ein = np.append(Ein, (x[i].dot(w) - y[i]))
    return Ein


## 四分法计算异常点
def calTukeyTest(y0):
    y = np.sort(y0)
    q1 = y[len(y) / 4 - 1]
    q3 = y[len(y) * 3 / 4 - 1]
    dv = q3 - q1
    high = q3 + 1.5 * dv
    low = q1 - 1.5 * dv
    print "high=", high
    print "low=", low
    print y0


##方差法寻找特殊点
def findException(data, w):
    x = data[:, 0]
    y = data[:, 1]
    x_ = np.zeros((len(data), 2))
    x_[:, 1] = data[:, 0]
    x_[:, 0] = np.ones(len(x))
    Eall = calE(x_, y, w)
    Ein = (np.sum(np.abs(Eall)) / len(data))
    out = np.array([])
    for index, e in enumerate(np.abs(Eall)):
        if e > 2 * Ein:
            out = np.append(out, index + 1)
###    print "方差法异常月份：", out
    return out.astype(np.int32)


## 判断总体趋势以及特殊点对趋势的影响(单个曲线)
def theEnd(data, s):
    x = data[:, 0]
    y = data[:, 1]
    # ## 折线图
    # plt.plot(x, y, marker='o', label="data")
    w = caseTrend(data, s)
    # plt.plot([1, len(data)], [w[0] + w[1], w[0] + w[1] * len(data)], label="trend")
    print s
    if w[1] > 0:
        print "总体呈上升趋势"
    else:
        print "总体呈下降趋势"
    excetPoint = findException(data, w)
    if len(excetPoint) > 0:
        for i in range(len(excetPoint)):
            point = excetPoint[i]
            data_ = np.append(data[:point - 1, :], data[point:, :], axis=0)
            w_ = caseTrend(data_, s + "_")
            # plt.plot([1, len(data)], [w_[0] + w_[1], w_[0] + w_[1] * len(data)], label=("trend_" + str(i)))
            # if w[1] > 0:
            if w_[1] > 0:
                if w[1] > w_[1]:
                    print "主要是%d月的影响使案件数量上升更明显" % point
                else:
                    print "因%d月的影响使案件数量上升变缓" % point
            else:
                print "因%d月的影响使案件数量从下降趋势变为了上升趋势" % point
        else:
            if w_[1] < 0:
                if w[1] < w_[1]:
                    print "主要是%d月的影响使案件数量下降更明显" % point
                else:
                    print "因%d月的影响使案件数量下降变缓" % point
            else:
                     print "因%d月的影响使案件数量从上升趋势变为了下降趋势" % point
    # plt.title(s)
    # plt.xlabel("month")
    # plt.ylabel("cases")
    # plt.legend()
    # plt.grid(True)
    # plt.show()


## 分析各个分趋势对总趋势的影响
def theEnd2(data, description, dataSum):
    s = "data"
    ws = []
    for index, item in enumerate(data):
        x = item[:, 0]
        y = item[:, 1]
        # plt.plot(x, y, marker='o', label=s + str(index))
        w = caseTrend(item, s + str(index))
        ws.append(w)
        logging.info('w%d = %f' % (index, w[1]))
    x = dataSum[:, 0]
    y = dataSum[:, 1]
    # plt.plot(x, y, marker='o', label=s + 'Sum')
    wSum = caseTrend(dataSum, s + 'Sum')
    logging.info('wSum = %f' % wSum[1])
    # plt.title(s)
    # plt.xlabel("month")
    # plt.ylabel("value")
    # plt.legend()
    # plt.grid(True)
    # plt.show()
    # for index, _ in enumerate(ws):
    #     plt.plot([1, len(data)], [_[1] + _[0], _[0] + _[1] * len(data)], label=s + str(index))
    # plt.plot([1, len(data)], [wSum[1] + wSum[0], wSum[0] + wSum[1] * len(data)], label="Sum")
    # plt.title("trend")
    # plt.xlabel("month")
    # plt.ylabel("value")
    # plt.legend()
    # plt.grid(True)
    # plt.show()
    ## 分析趋势
    ws = np.array(ws)
    wout = []
    sout = ""
    temp = 0
    wSum = wSum[1]
    for index, _ in enumerate(ws):
        if wSum > 0 and _[1] > wSum:
            wout.append(index)
            if temp == 1: sout += "和"
            sout += description[index]
            temp = 1
        if wSum < 0 and _[1] < wSum:
            wout.append(index)
            if temp == 1: sout += "和"
            sout += description[index]
            temp = 1
    if sout == "": sout = "所有因素共同结果"
    if wSum > 0:
        print "总结：影响%s上升趋势的主要是\"%s\"" % (description[-1], sout)
    else:
       print "总结：影响%s下降趋势的主要是\"%s\"" % (description[-1], sout)


##查找xls表格int数据(行)
def selectintValues(book, sheet, d1, d2):
    sheet = book[sheet]
    s = sheet[d1:d2]
    values = []
    rowvalues = []
    for row in s:
        for i in row:
            rowvalues.append(int(i.value))
        values.append(rowvalues)
        rowvalues = []
    data = np.array(values)
    #     print data
    return data


##查找xls表格str数据(列)
def selectstrValues(book, sheet, d1, d2):
    sheet = book[sheet]
    s = sheet[d1:d2]
    values = []
    for row in s:
        for i in row:
            values.append((i.value).encode('utf-8'))
    description = values
    #     print description
    return description


def selectDataFromXls():
    book = load_workbook("data.xlsx", data_only=True)
    selectintValues(book, 'Sheet1', 'd5', 'o9')
    selectstrValues(book, 'Sheet1', 'c5', 'c9')


def selectDataFromMysql():
    import mysql.connector
    conn = mysql.connector.connect(user='root', password='root', database='111', use_unicode=True)
    cursor = conn.cursor()
    cursor.execute('select * from Data ')
    # cursor.execute('select * from user where id = %s', ('1',))
    values = cursor.fetchall()
    cursor.close()
    conn.close()
    print values


def selectintValues1(book, sheet, d1, d2):
    data = selectintValues(book, sheet, d1, d2)[0]
    datao = np.zeros((len(data), 2))
    datao[:, 0] = range(1, len(data) + 1)
    datao[:, 1] = data
    return datao.astype(int)


def selectintValues2(book, sheet, d1, d2):
    def abc(data):
        datao = np.zeros((len(data), 2))
        datao[:, 0] = range(1, len(data) + 1)
        datao[:, 1] = data
        return datao.astype(int)

    data = selectintValues(book, sheet, d1, d2)
    datasum = []
    datas = []
    for index, item in enumerate(data):
        if index == 0:
            datasum = abc(item)
        else:
            datas.append(abc(item))
    return np.array(datasum), datas


def q1():
    book = load_workbook("data.xlsx", data_only=True)
    data1 = selectintValues1(book, 'Sheet1', 'd5', 'o5')
    data2 = selectintValues1(book, 'Sheet1', 'd19', 'o19')
    data3 = selectintValues1(book, 'Sheet1', 'd34', 'o34')
    #     data1 = np.array([[1,3421],[2,4199],[3,3612],[4,3569],[5,3525],[6,3459],[7,3405],[8,3347],[9,3229],[10,1208],[11,4187],[12,4155]])
    theEnd(data1, "data1")


# #     data2 = np.array([[1,3087],[2,4555],[3,7815],[4,9705],[5,4700],[6,4703],[7,5252],[8,4709],[9,4712],[10,1705],[11,4718],[12,4721]])
#     theEnd(data2,"data2")
# #     data3 = np.array([[1,1538],[2,1681],[3,1961],[4,1741],[5,1448],[6,1748],[7,1784],[8,2036],[9,1732],[10,1182],[11,1797],[12,1744]])
#     theEnd(data3,"data3")

def q2():
    # book = load_workbook("data.xlsx", data_only=True)
    # dataSum, data = selectintValues2(book, 'Sheet1', 'd5', 'o9')
    # description0 = selectstrValues(book, 'Sheet1', 'c5', 'c9')
    # #     print "dataSum:",dataSum
    # description = description0[1:]
    # description.append(description0[0])

    data0 = np.array([[1,600],[2,601],[3,400],[4,603],[5,345],[6,605],[7,606],[8,567],[9,608],[10,643],[11,610],[12,611]])
    data1 = np.array([[1,400],[2,478],[3,893],[4,403],[5,343],[6,405],[7,406],[8,407],[9,408],[10,19],[11,410],[12,411]])
    data2 = np.array([[1,169],[2,189],[3,202],[4,232],[5,256],[6,205],[7,266],[8,555],[9,208],[10,11],[11,267],[12,211]])
    data3 = np.array([[1,369],[2,413],[3,466],[4,503],[5,504],[6,533],[7,506],[8,507],[9,508],[10,509],[11,510],[12,511]])
    data = [data0,data1,data2,data3]
    # print data
    description = ('两强一盗','酒驾','凶杀','其他','案件总数')
    dataSum = data0+data1+data2+data3
    dataSum[:,0] = dataSum[:,0]/(len(data))
    theEnd2(data, description, dataSum)


if __name__ == '__main__':
    start = time.time()
    for i in range(500):
        q1()
        q2()
    elapsed = (time.time() - start)
    print "time:",elapsed
